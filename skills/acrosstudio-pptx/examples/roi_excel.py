"""
01-04 ROI 試算 Excel v2
- 装置数を前提条件で変更可能
- 運用費を今回作成した運用費スライドと整合
- クラウド / オンプレを選択可能
- 4 テーマ独立で ROI 試算
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from copy import copy

# 既存ファイルをベースにする（なければ新規作成）
# 注：この例は「既存の ROI Excel をスタイル整形する」用途を想定している。
#     既存ファイルがある場合は src に渡す。無い場合は空ブックから作る。
src = '/mnt/user-data/outputs/01-04_ROI_試算.xlsx'
dst = '/home/claude/work/01-04_ROI_試算_v2.xlsx'

import os
if os.path.exists(src):
    wb = openpyxl.load_workbook(src)
else:
    # ベースファイルが無い場合は、スクリプトが前提とするシート構成を最小限用意する。
    # 注：この例は本来「既存の ROI Excel を Acrosstudio スタイルに整形する」用途。
    #     新規から作る場合は、前提条件シートの値を実データで埋め直すこと。
    wb = openpyxl.Workbook()
    base = wb.active
    base.title = '前提条件'
    # スクリプトが参照する既存セル（R6 評価期間 / R7 人件費単価 / R8 年間営業日数）の枠だけ用意
    base['B6'] = '評価期間（年）';     base['C6'] = 5
    base['B7'] = '人件費単価（円/h）'; base['C7'] = 5000
    base['B8'] = '年間営業日数（日）'; base['C8'] = 240

# ─── スタイル定義 ───
NAVY = "073763"
NAVY_LIGHT = "DCE6F0"
MAGENTA = "A22B94"
MAGENTA_LIGHT = "F4E4F0"
INPUT_BG = "FFF2CC"  # 入力セル（黄色背景）
INPUT_FONT = "0070C0"  # 入力セル（青字）
CALC_BG = "F2F2F2"   # 計算セル（グレー背景）
RESULT_BG = "073763"  # 結果セル（濃紺背景）
RESULT_FG = "FFFFFF"  # 結果セル（白字）

thin = Side(border_style="thin", color="999999")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_section_header(cell):
    cell.font = Font(bold=True, size=12, color=NAVY)

def style_input(cell):
    cell.fill = PatternFill("solid", fgColor=INPUT_BG)
    cell.font = Font(color=INPUT_FONT, bold=True)
    cell.border = border_all

def style_calc(cell):
    cell.fill = PatternFill("solid", fgColor=CALC_BG)
    cell.border = border_all

def style_result(cell):
    cell.fill = PatternFill("solid", fgColor=RESULT_BG)
    cell.font = Font(color=RESULT_FG, bold=True, size=12)
    cell.border = border_all

def style_label(cell, bold=False):
    cell.font = Font(bold=bold)
    cell.border = border_all

# ─── 前提条件シートを更新 ───
ws_pre = wb['前提条件']

# 共通パラメータに「対象装置数」「保管期間」 を追加
# 既存：R6 評価期間, R7 人件費単価, R8 年間営業日数
# 追加：R9 対象装置数, R10 保管期間（情報用）

ws_pre.cell(row=9, column=2, value="対象装置数（台）")
ws_pre.cell(row=9, column=3, value=50)
ws_pre.cell(row=9, column=5, value="客先で稼働中の半導体洗浄装置の台数。変更すると効果値・運用費が連動。")
style_input(ws_pre.cell(row=9, column=3))
style_label(ws_pre.cell(row=9, column=2))

ws_pre.cell(row=10, column=2, value="データ保管期間（年）")
ws_pre.cell(row=10, column=3, value=5)
ws_pre.cell(row=10, column=5, value="運用費の試算で使用。長期保管ほど高額。")
style_input(ws_pre.cell(row=10, column=3))
style_label(ws_pre.cell(row=10, column=2))

# ROI 計算の考え方も更新（R12-R17 だが、説明追加）
# R18 に新しい説明追加
ws_pre.cell(row=18, column=2, value="・年間運用保守は内訳形式（インフラ + データ転送 + UI・監視・セキュリティ等）")
ws_pre.cell(row=19, column=2, value="・構成タイプ（クラウド / オンプレ）は各テーマシートで選択")
ws_pre.cell(row=20, column=2, value="・装置数を変更すると、データ蓄積費・推論費・効果値が連動")


# ─── 各テーマシートを更新 ───
# 各テーマで運用費の内訳と装置数連動を実装

# 各テーマの運用費パラメータ（50 台基準、月額）
theme_params = {
    "01_Predictive": {
        "title": "01  故障予知保全（Predictive）",
        # クラウド構成の月額（50 台基準）
        "cloud": {
            "データ蓄積（時系列 DB・オブジェクトストレージ）": ("装置数連動", 20000),  # 月 2 万
            "推論（劣化予兆検知、バッチ集約）": ("装置数連動", 35000),  # 月 3.5 万
            "データ転送（拠点間 VPN）": ("装置数連動", 15000),  # 月 1.5 万
            "UI・監視・セキュリティ機能": ("固定", 40000),  # 月 4 万
        },
        "onprem": {
            "サーバ電気代（推論サーバ 2〜3 台）": ("装置数連動", 2000),  # 月 2 千円
            "ストレージ電気代・バックアップ": ("固定", 15000),  # 月 1.5 万
            "データセンター利用料（ハーフラック相当）": ("固定", 40000),  # 月 4 万
            "拠点間 VPN・セキュリティ機能": ("固定", 20000),  # 月 2 万
        },
        # 効果項目（既存を装置数連動に修正）
        "effects": [
            # ラベル, 現状値（50 台基準）, 単位, 改善率, 装置数連動か, 計算式タイプ, 計算根拠
            ("計画外停止の削減", 12, "回 × 円/回", 0.4, True, "回数×単価×改善率", "ベローズポンプ起因の停止 × 1 回損失 3,000,000 円 × 削減率 40%"),
            ("緊急対応コストの削減", 5000000, "円/年", 0.7, True, "金額×改善率", "年間緊急対応費用 × 計画化率 70%"),
            ("装置寿命の延長効果", 2000000, "円/年", None, True, "金額", "予兆検知による部品の劣化前交換で装置寿命が延びる効果"),
        ],
        # 効果項目の特殊計算（01 Predictive のみ）
        "effects_formula": {
            0: "=C{row}*3000000*E{row}",  # 計画外停止
        },
    },
    "02_ProcessAdj": {
        "title": "02  プロセス調整（Process Adjustment）",
        "cloud": {
            "データ蓄積（プロセス変数・履歴）": ("装置数連動", 20000),
            "推論（影響係数モデル適用、バッチ集約）": ("装置数連動", 55000),
            "データ転送（拠点間 VPN・03 連携）": ("装置数連動", 15000),
            "Web UI・監視・セキュリティ機能": ("固定", 40000),
        },
        "onprem": {
            "サーバ電気代（推論 + Web 3〜4 台）": ("装置数連動", 4000),
            "ストレージ電気代・バックアップ": ("固定", 15000),
            "データセンター利用料": ("固定", 55000),
            "拠点間 VPN・セキュリティ機能": ("固定", 20000),
        },
        "effects": [
            ("品質ばらつき低減 → 歩留まり向上", 50000000, "円/年", 0.15, True, "金額×改善率", "ばらつきによる年間不良ロット損失 × 削減率 15%"),
            ("装置間品質差の縮小", 3000000, "円/年", 0.3, True, "金額×改善率", "顧客クレーム対応・装置調整出張等 × 削減率 30%"),
            ("担当者の調整作業工数削減", 1000, "時間 × 円/時", 0.3, False, "時間×単価×改善率", "年間調整作業工数 × 人件費単価 × 削減率 30%"),
        ],
        "effects_formula": {
            2: "=C{row}*前提条件!$C$7*E{row}",  # 工数削減（人件費単価参照）
        },
    },
    "03_Development": {
        "title": "03  洗浄品質評価・要因分析（Development）",
        "cloud": {
            "画像ストレージ・バックアップ": ("装置数連動", 70000),
            "画像解析（GPU 推論、夜間バッチ集約）": ("装置数連動", 100000),
            "データ転送（大量画像、拠点間 VPN）": ("装置数連動", 35000),
            "UI・監視・セキュリティ機能": ("固定", 40000),
        },
        "onprem": {
            "GPU サーバ電気代（画像処理用）": ("固定", 15000),
            "高速ストレージ電気代・バックアップ": ("装置数連動", 30000),
            "データセンター利用料": ("固定", 75000),
            "拠点間 VPN・セキュリティ機能": ("固定", 20000),
        },
        "effects": [
            ("目視評価の自動化", 2000, "時間 × 円/時", 0.5, True, "時間×単価×改善率", "年間目視評価工数 × 人件費単価 × 自動化率 50%"),
            ("要因分析の高速化", 800, "時間 × 円/時", 0.4, True, "時間×単価×改善率", "年間要因分析工数 × 人件費単価 × 短縮率 40%"),
            ("良条件再現による不良率低減", 50000000, "円/年", 0.05, True, "金額×改善率", "年間不良ロット損失 × 削減率 5%"),
        ],
        "effects_formula": {
            0: "=C{row}*前提条件!$C$7*E{row}",
            1: "=C{row}*前提条件!$C$7*E{row}",
        },
    },
    "04_Design": {
        "title": "04  設計支援（Design）",
        "cloud": {
            "ベクトル DB・ドキュメントストレージ": ("固定", 45000),
            "LLM API 利用料（OpenAI / Anthropic 等）": ("固定", 120000),
            "データ転送・Web UI・監視": ("固定", 40000),
            "セキュリティ機能（機密文書扱い）": ("固定", 30000),
        },
        "onprem": {
            "サーバ電気代・データセンター利用料": ("固定", 45000),
            "LLM API 利用料（オンプレでも外部 API）": ("固定", 120000),
            "バックアップ・拠点間 VPN": ("固定", 30000),
            "セキュリティ機能（機密文書扱い）": ("固定", 20000),
        },
        "effects": [
            ("設計ミスの事前検知 → リワーク削減", 1500, "時間 × 円/時", 0.4, False, "時間×単価×改善率", "年間リワーク工数 × 人件費単価 × 検知率 40%"),
            ("不具合原因特定の高速化", 500, "時間 × 円/時", 0.5, False, "時間×単価×改善率", "年間原因特定工数 × 人件費単価 × 短縮率 50%"),
            ("若手の独立度向上 → 教育工数削減", 1000, "時間 × 円/時", 0.3, False, "時間×単価×改善率", "ベテラン教育工数 × 人件費単価 × 削減率 30%"),
        ],
        "effects_formula": {
            0: "=C{row}*前提条件!$C$7*E{row}",
            1: "=C{row}*前提条件!$C$7*E{row}",
            2: "=C{row}*前提条件!$C$7*E{row}",
        },
    },
}


def rebuild_theme_sheet(wb, sname, params):
    """各テーマシートを内訳形式 + 装置数連動で書き換え"""
    # 既存シートを削除して新規作成
    if sname in wb.sheetnames:
        idx = wb.sheetnames.index(sname)
        del wb[sname]
        ws = wb.create_sheet(sname, idx)
    else:
        ws = wb.create_sheet(sname)

    # 列幅
    widths = [2, 38, 18, 14, 12, 18, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # タイトル
    ws.cell(row=2, column=2, value=params["title"]).font = Font(bold=True, size=14, color=NAVY)

    # ─── 投資コスト ───
    ws.cell(row=5, column=2, value="■ 投資コスト（円・税抜）")
    style_section_header(ws.cell(row=5, column=2))

    headers = ["費用項目", "金額（円）", "区分", "", "", "備考"]
    for i, h in enumerate(headers):
        c = ws.cell(row=6, column=2+i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border = border_all

    invest_items = [
        ("要求整理・提案", 300000, "進行中", "提案作成活動。別途見積。"),
        ("要件定義", 5000000, "本提案範囲", "約 2 ヶ月。仕様と最終金額を確定。"),
        ("開発", 20000000, "概算範囲", "約 3〜4 ヶ月。要件定義完了後に正式見積。"),
        ("検証（実環境）", 10000000, "後続", "連携・実環境検証。"),
    ]
    for i, (label, amount, kind, note) in enumerate(invest_items):
        r = 7 + i
        ws.cell(row=r, column=2, value=label)
        c_amount = ws.cell(row=r, column=3, value=amount)
        c_amount.number_format = '#,##0'
        style_input(c_amount)
        ws.cell(row=r, column=4, value=kind)
        ws.cell(row=r, column=7, value=note)
        for col in [2, 4, 7]:
            ws.cell(row=r, column=col).border = border_all
            ws.cell(row=r, column=col).font = Font()

    # 初期投資合計
    ws.cell(row=11, column=2, value="初期投資 合計").font = Font(bold=True)
    c_total = ws.cell(row=11, column=3, value="=SUM(C7:C10)")
    c_total.number_format = '#,##0'
    style_calc(c_total)
    c_total.font = Font(bold=True)
    ws.cell(row=11, column=7, value="要求整理 + 要件定義 + 開発 + 検証")
    for col in [2, 4, 7]:
        ws.cell(row=11, column=col).border = border_all

    # ─── 年間運用保守（内訳形式、装置数連動）───
    ws.cell(row=14, column=2, value="■ 年間運用保守（インフラ費）").font = Font(bold=True, size=12, color=NAVY)

    # 構成タイプ選択
    ws.cell(row=15, column=2, value="構成タイプ（cloud / onprem）").font = Font(bold=True)
    c_type = ws.cell(row=15, column=3, value="cloud")
    style_input(c_type)
    ws.cell(row=15, column=7, value="cloud or onprem を入力。運用費の数式が切り替わる。")
    ws.cell(row=15, column=2).border = border_all
    ws.cell(row=15, column=7).border = border_all

    # 対象装置数（前提条件から）
    ws.cell(row=16, column=2, value="対象装置数（前提条件から参照）").font = Font(bold=True)
    c_eqcount = ws.cell(row=16, column=3, value="=前提条件!C9")
    style_calc(c_eqcount)
    c_eqcount.font = Font(color="2D7D2D")  # 緑字（シート参照）
    ws.cell(row=16, column=7, value="前提条件シートで変更してください")
    ws.cell(row=16, column=2).border = border_all
    ws.cell(row=16, column=7).border = border_all

    # 内訳ヘッダ
    sub_headers = ["内訳項目", "月額（円）", "クラウド月額", "オンプレ月額", "", "計算根拠"]
    for i, h in enumerate(sub_headers):
        c = ws.cell(row=17, column=2+i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border = border_all

    # クラウド・オンプレの内訳を表示
    cloud_items = list(params["cloud"].items())
    onprem_items = list(params["onprem"].items())

    # 装置数 50 台基準で、装置数連動 / 固定 を切り替え
    for i in range(4):
        r = 18 + i
        cloud_label, (cloud_mode, cloud_amount) = cloud_items[i]
        onprem_label, (onprem_mode, onprem_amount) = onprem_items[i]

        # 内訳項目名：構成タイプに応じて切り替え
        # シンプル：クラウド月額とオンプレ月額を両方表示、合計は IF で切り替え
        ws.cell(row=r, column=2, value=f"クラウド: {cloud_label}\nオンプレ: {onprem_label}")
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

        # クラウド月額（装置数連動 or 固定）
        if cloud_mode == "装置数連動":
            cloud_formula = f"={cloud_amount}*C16/50"
        else:
            cloud_formula = f"={cloud_amount}"
        c_cloud = ws.cell(row=r, column=4, value=cloud_formula)
        c_cloud.number_format = '#,##0'
        style_calc(c_cloud)

        # オンプレ月額（装置数連動 or 固定）
        if onprem_mode == "装置数連動":
            onprem_formula = f"={onprem_amount}*C16/50"
        else:
            onprem_formula = f"={onprem_amount}"
        c_onprem = ws.cell(row=r, column=5, value=onprem_formula)
        c_onprem.number_format = '#,##0'
        style_calc(c_onprem)

        # 採用月額（構成タイプに応じて IF で切り替え）
        c_adopt = ws.cell(row=r, column=3, value=f'=IF($C$15="cloud",D{r},E{r})')
        c_adopt.number_format = '#,##0'
        c_adopt.fill = PatternFill("solid", fgColor=NAVY_LIGHT)
        c_adopt.font = Font(bold=True)
        c_adopt.border = border_all

        # 計算根拠
        mode_text = "装置数連動（基準 50 台）" if cloud_mode == "装置数連動" or onprem_mode == "装置数連動" else "固定費"
        ws.cell(row=r, column=7, value=mode_text)
        ws.cell(row=r, column=2).border = border_all
        ws.cell(row=r, column=7).border = border_all
        ws.row_dimensions[r].height = 30

    # 月額合計
    r_month = 22
    ws.cell(row=r_month, column=2, value="月額 合計").font = Font(bold=True)
    c_msum = ws.cell(row=r_month, column=3, value=f"=SUM(C18:C21)")
    c_msum.number_format = '#,##0'
    c_msum.fill = PatternFill("solid", fgColor=NAVY_LIGHT)
    c_msum.font = Font(bold=True, size=12)
    c_msum.border = border_all
    ws.cell(row=r_month, column=4, value=f"=SUM(D18:D21)").number_format = '#,##0'
    ws.cell(row=r_month, column=4).border = border_all
    ws.cell(row=r_month, column=5, value=f"=SUM(E18:E21)").number_format = '#,##0'
    ws.cell(row=r_month, column=5).border = border_all
    ws.cell(row=r_month, column=2).border = border_all

    # 年額合計（運用保守の値）
    r_year = 23
    ws.cell(row=r_year, column=2, value="年額 合計（採用構成）").font = Font(bold=True)
    c_ysum = ws.cell(row=r_year, column=3, value=f"=C{r_month}*12")
    c_ysum.number_format = '#,##0'
    style_result(c_ysum)
    ws.cell(row=r_year, column=7, value="年間運用保守として ROI 計算に使用")
    ws.cell(row=r_year, column=2).border = border_all
    ws.cell(row=r_year, column=7).border = border_all

    # ─── 効果項目 ───
    r_eff_header = 26
    ws.cell(row=r_eff_header, column=2, value="■ 効果項目（年間）").font = Font(bold=True, size=12, color=NAVY)

    eff_headers = ["効果項目", "現状値（50 台基準）", "単位", "改善率", "年間効果（円）", "計算根拠"]
    for i, h in enumerate(eff_headers):
        c = ws.cell(row=r_eff_header+1, column=2+i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border = border_all

    effects = params["effects"]
    effects_formula = params.get("effects_formula", {})
    eff_start = r_eff_header + 2  # 28
    for i, (label, base_val, unit, rate, eqlinked, formula_type, basis) in enumerate(effects):
        r = eff_start + i
        ws.cell(row=r, column=2, value=label)
        # 現状値（装置数連動の場合は数式に）
        if eqlinked:
            c_val = ws.cell(row=r, column=3, value=f"={base_val}*C16/50")
        else:
            c_val = ws.cell(row=r, column=3, value=base_val)
        c_val.number_format = '#,##0'
        style_input(c_val) if not eqlinked else style_calc(c_val)
        ws.cell(row=r, column=4, value=unit)
        if rate is not None:
            c_rate = ws.cell(row=r, column=5, value=rate)
            c_rate.number_format = '0%'
            style_input(c_rate)
        else:
            ws.cell(row=r, column=5, value="—")

        # 年間効果の計算式
        if i in effects_formula:
            ce_formula = effects_formula[i].format(row=r)
        elif rate is not None:
            ce_formula = f"=C{r}*E{r}"  # 金額×改善率
        else:
            ce_formula = f"=C{r}"  # そのまま
        c_eff = ws.cell(row=r, column=6, value=ce_formula)
        c_eff.number_format = '#,##0'
        style_calc(c_eff)

        ws.cell(row=r, column=7, value=basis).alignment = Alignment(wrap_text=True, vertical="top")
        for col in [2, 4, 7]:
            ws.cell(row=r, column=col).border = border_all

    r_eff_total = eff_start + len(effects)  # 31
    ws.cell(row=r_eff_total, column=2, value="年間効果 合計").font = Font(bold=True)
    c_efftot = ws.cell(row=r_eff_total, column=6, value=f"=SUM(F{eff_start}:F{r_eff_total-1})")
    c_efftot.number_format = '#,##0'
    c_efftot.fill = PatternFill("solid", fgColor=NAVY_LIGHT)
    c_efftot.font = Font(bold=True, size=12)
    c_efftot.border = border_all
    ws.cell(row=r_eff_total, column=2).border = border_all

    # ─── ROI 試算 ───
    r_roi_header = r_eff_total + 3  # 34
    ws.cell(row=r_roi_header, column=2, value="■ ROI 試算").font = Font(bold=True, size=12, color=NAVY)

    roi_headers = ["計算項目", "値", "計算式"]
    for i, h in enumerate(roi_headers):
        c = ws.cell(row=r_roi_header+1, column=2+i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border = border_all

    roi_start = r_roi_header + 2  # 36
    # 評価期間
    ws.cell(row=roi_start, column=2, value="評価期間")
    c = ws.cell(row=roi_start, column=3, value=f"=前提条件!C6")
    c.number_format = '0"年"'
    style_calc(c)
    c.font = Font(color="2D7D2D")
    ws.cell(row=roi_start, column=4, value="前提条件シートから参照")

    # 投資総額
    ws.cell(row=roi_start+1, column=2, value="投資総額（評価期間中）")
    c = ws.cell(row=roi_start+1, column=3, value=f"=C11+C23*C{roi_start}")
    c.number_format = '#,##0'
    style_calc(c)
    ws.cell(row=roi_start+1, column=4, value="初期投資 + 年額運用保守 × 評価期間")

    # 年間ネット効果
    ws.cell(row=roi_start+2, column=2, value="年間ネット効果")
    c = ws.cell(row=roi_start+2, column=3, value=f"=F{r_eff_total}-C23")
    c.number_format = '#,##0'
    style_calc(c)
    ws.cell(row=roi_start+2, column=4, value="年間効果 − 年額運用保守")

    # 累計ネット効果
    ws.cell(row=roi_start+3, column=2, value="累計ネット効果（評価期間）")
    c = ws.cell(row=roi_start+3, column=3, value=f"=C{roi_start+2}*C{roi_start}")
    c.number_format = '#,##0'
    style_calc(c)
    ws.cell(row=roi_start+3, column=4, value="年間ネット効果 × 評価期間")

    # 純利益
    ws.cell(row=roi_start+4, column=2, value="純利益（評価期間）")
    c = ws.cell(row=roi_start+4, column=3, value=f"=C{roi_start+3}-C11")
    c.number_format = '#,##0'
    style_calc(c)
    ws.cell(row=roi_start+4, column=4, value="累計ネット効果 − 初期投資")

    # ROI
    ws.cell(row=roi_start+6, column=2, value="ROI（%）").font = Font(bold=True)
    c = ws.cell(row=roi_start+6, column=3, value=f"=IF(C11=0,0,C{roi_start+4}/C11)")
    c.number_format = '0.0%'
    style_result(c)
    ws.cell(row=roi_start+6, column=4, value="純利益 / 初期投資")

    # 回収期間
    ws.cell(row=roi_start+7, column=2, value="投資回収期間（年）").font = Font(bold=True)
    c = ws.cell(row=roi_start+7, column=3, value=f"=IF(C{roi_start+2}<=0,\"効果不足\",C11/C{roi_start+2})")
    c.number_format = '0.0"年"'
    style_result(c)
    ws.cell(row=roi_start+7, column=4, value="初期投資 / 年間ネット効果")

    # ボーダーを ROI セクションに付与
    for r in range(roi_start, roi_start+8):
        for col in [2, 4]:
            ws.cell(row=r, column=col).border = border_all


# 各テーマシートを再構築
for sname, params in theme_params.items():
    rebuild_theme_sheet(wb, sname, params)


# ─── サマリシートを再構築 ───
# 結合セルなどがあるので、既存を削除して新規作成
if 'サマリ' in wb.sheetnames:
    idx = wb.sheetnames.index('サマリ')
    del wb['サマリ']
    ws_sum = wb.create_sheet('サマリ', idx)
else:
    ws_sum = wb.create_sheet('サマリ', 0)

# 列幅
sum_widths = [2, 30, 18, 18, 18, 12, 14]
for i, w in enumerate(sum_widths, 1):
    ws_sum.column_dimensions[get_column_letter(i)].width = w

# タイトル
ws_sum.cell(row=2, column=2, value="半導体洗浄装置 AI 活用 4 テーマ  ROI サマリ（サンプル）").font = Font(bold=True, size=14, color=NAVY)
ws_sum.cell(row=3, column=2, value="各テーマシートで入力された投資・効果・運用費に基づく ROI 試算結果を一覧します。").font = Font(size=10, color="595959")
ws_sum.cell(row=4, column=2, value="装置数・構成タイプを各シートで変更すると、本サマリも自動更新されます。").font = Font(size=10, color="595959")

# テーマ別 ROI 一覧

themes_summary = [
    ("01_Predictive", "01 故障予知保全", 7),
    ("02_ProcessAdj", "02 プロセス調整提案", 8),
    ("03_Development", "03 洗浄品質評価・要因分析", 9),
    ("04_Design", "04 設計支援", 10),
]

# サマリ表のヘッダ
ws_sum.cell(row=6, column=2, value="テーマ").font = Font(bold=True, color="FFFFFF")
ws_sum.cell(row=6, column=3, value="初期投資（円）").font = Font(bold=True, color="FFFFFF")
ws_sum.cell(row=6, column=4, value="年間効果（円）").font = Font(bold=True, color="FFFFFF")
ws_sum.cell(row=6, column=5, value="年間運用保守（円）").font = Font(bold=True, color="FFFFFF")
ws_sum.cell(row=6, column=6, value="ROI（%）").font = Font(bold=True, color="FFFFFF")
ws_sum.cell(row=6, column=7, value="回収期間").font = Font(bold=True, color="FFFFFF")
for i in range(2, 8):
    ws_sum.cell(row=6, column=i).fill = PatternFill("solid", fgColor=NAVY)
    ws_sum.cell(row=6, column=i).alignment = Alignment(horizontal="center")
    ws_sum.cell(row=6, column=i).border = border_all

# サマリのデータ行
# 新しい構造での行参照：
#  - 初期投資合計：C11
#  - 年額運用保守：C23
#  - 年間効果合計：F31
#  - ROI：C42 (roi_start+6 = 36+6 = 42)
#  - 回収期間：C43 (roi_start+7 = 43)
for sname, label, row in themes_summary:
    ws_sum.cell(row=row, column=2, value=label)
    ws_sum.cell(row=row, column=3, value=f"='{sname}'!C11").number_format = '#,##0'
    ws_sum.cell(row=row, column=4, value=f"='{sname}'!F31").number_format = '#,##0'
    ws_sum.cell(row=row, column=5, value=f"='{sname}'!C23").number_format = '#,##0'
    ws_sum.cell(row=row, column=6, value=f"='{sname}'!C42").number_format = '0.0%'
    ws_sum.cell(row=row, column=7, value=f"='{sname}'!C43").number_format = '0.0"年"'
    for col in range(2, 8):
        ws_sum.cell(row=row, column=col).border = border_all
        ws_sum.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F2F2F2")
    ws_sum.cell(row=row, column=2).font = Font(bold=True)

# 合計行
ws_sum.cell(row=11, column=2, value="合計（4 テーマ）").font = Font(bold=True)
ws_sum.cell(row=11, column=3, value=f"=SUM(C7:C10)").number_format = '#,##0'
ws_sum.cell(row=11, column=4, value=f"=SUM(D7:D10)").number_format = '#,##0'
ws_sum.cell(row=11, column=5, value=f"=SUM(E7:E10)").number_format = '#,##0'
for col in [2, 3, 4, 5]:
    ws_sum.cell(row=11, column=col).fill = PatternFill("solid", fgColor=NAVY_LIGHT)
    ws_sum.cell(row=11, column=col).font = Font(bold=True)
    ws_sum.cell(row=11, column=col).border = border_all
for col in [6, 7]:
    ws_sum.cell(row=11, column=col).border = border_all
    ws_sum.cell(row=11, column=col).fill = PatternFill("solid", fgColor=NAVY_LIGHT)

# 注記更新
ws_sum.cell(row=13, column=2, value="■ 留意事項").font = Font(bold=True, size=12, color=NAVY)
notes = [
    "※ 本シートの数値はすべて参考値（仮置き）です。各テーマシートの入力欄を貴社実態に応じて変更すると、サマリも自動更新されます。",
    "※ 装置数は前提条件シートで変更可能。装置数を変えると、効果値・運用保守費が連動して変動します。",
    "※ 構成タイプ（cloud / onprem）は各テーマシートで個別選択。運用保守の計算が切り替わります。",
    "※ 01〜04 はそれぞれ独立した試算で、4 テーマを必ずすべて実施する前提ではありません。テーマごとに採否をご判断ください。",
    "※ 貴社（装置メーカー）の直接効果と、貴社顧客（半導体メーカー）への提供価値による間接効果を含めて評価しています。",
    "※ 運用保守は本資料の運用費試算スライドと整合する内訳構成。問合せ対応・ML 再学習は含まず、必要時に別途。",
]
for i, n in enumerate(notes):
    ws_sum.cell(row=14+i, column=2, value=n).font = Font(size=10, color="595959")

# 保存
wb.save(dst)
print(f"完了: {dst}")
